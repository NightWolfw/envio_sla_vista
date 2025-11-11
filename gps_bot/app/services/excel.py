"""
Serviço de importação e exportação de planilhas Excel
Funções para importar grupos do Excel e exportar para Excel
"""
import openpyxl
from io import BytesIO
from app.models.database import get_db_site


def importar_excel(filepath):
    """
    Importa grupos de planilha Excel para o banco de dados

    Args:
        filepath: Caminho do arquivo Excel (.xlsx)

    Returns:
        dict: Estatísticas da importação (importados, atualizados, erros)
    """
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        conn = get_db_site()
        cur = conn.cursor()

        importados = 0
        atualizados = 0
        erros = 0

        # Ignora cabeçalho (linha 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                # Espera colunas: group_id, nome_grupo, envio, cr, cliente, pec_01, pec_02,
                # diretorexecutivo, diretorregional, gerenteregional, gerente, supervisor, dia_todo
                group_id = str(row[0]).strip() if row[0] else None
                nome_grupo = str(row[1]).strip() if row[1] else None
                envio = bool(row[2]) if row[2] is not None else True
                cr = str(row[3]).strip() if row[3] else None
                cliente = str(row[4]).strip() if row[4] else None
                pec_01 = str(row[5]).strip() if row[5] else None
                pec_02 = str(row[6]).strip() if row[6] else None
                diretor_executivo = str(row[7]).strip() if row[7] else None
                diretor_regional = str(row[8]).strip() if row[8] else None
                gerente_regional = str(row[9]).strip() if row[9] else None
                gerente = str(row[10]).strip() if row[10] else None
                supervisor = str(row[11]).strip() if row[11] else None
                dia_todo = bool(row[12]) if row[12] is not None else False

                if not group_id or not nome_grupo:
                    erros += 1
                    continue

                # Verifica se já existe
                cur.execute("SELECT id FROM grupos_whatsapp WHERE group_id = %s", (group_id,))
                existe = cur.fetchone()

                if existe:
                    # Atualiza
                    cur.execute("""
                        UPDATE grupos_whatsapp 
                        SET nome_grupo = %s, envio = %s, cr = %s, cliente = %s,
                            pec_01 = %s, pec_02 = %s, diretorexecutivo = %s,
                            diretorregional = %s, gerenteregional = %s, gerente = %s,
                            supervisor = %s, dia_todo = %s
                        WHERE group_id = %s
                    """, (nome_grupo, envio, cr, cliente, pec_01, pec_02,
                          diretor_executivo, diretor_regional, gerente_regional,
                          gerente, supervisor, dia_todo, group_id))
                    atualizados += 1
                else:
                    # Insere novo
                    cur.execute("""
                        INSERT INTO grupos_whatsapp 
                        (group_id, nome_grupo, envio, cr, cliente, pec_01, pec_02,
                         diretorexecutivo, diretorregional, gerenteregional, gerente,
                         supervisor, dia_todo)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (group_id, nome_grupo, envio, cr, cliente, pec_01, pec_02,
                          diretor_executivo, diretor_regional, gerente_regional,
                          gerente, supervisor, dia_todo))
                    importados += 1

            except Exception as e:
                print(f"Erro ao importar linha: {str(e)}")
                erros += 1
                continue

        conn.commit()
        cur.close()
        conn.close()

        print(f"✅ Importação concluída: {importados} novos, {atualizados} atualizados, {erros} erros")

        return {
            'importados': importados,
            'atualizados': atualizados,
            'erros': erros
        }

    except Exception as e:
        print(f"❌ Erro na importação: {str(e)}")
        raise


def exportar_excel():
    """
    Exporta grupos do banco de dados para planilha Excel

    Returns:
        BytesIO: Buffer com arquivo Excel em memória
    """
    try:
        conn = get_db_site()
        cur = conn.cursor()

        cur.execute("""
            SELECT group_id, nome_grupo, envio, cr, cliente, pec_01, pec_02,
                   diretorexecutivo, diretorregional, gerenteregional, gerente,
                   supervisor, dia_todo
            FROM grupos_whatsapp
            ORDER BY nome_grupo
        """)

        grupos = cur.fetchall()
        cur.close()
        conn.close()

        # Cria nova planilha
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grupos WhatsApp"

        # Cabeçalhos
        headers = [
            'ID Grupo', 'Nome do Grupo', 'Envio Ativo', 'CR', 'Cliente',
            'PEC 01', 'PEC 02', 'Diretor Executivo', 'Diretor Regional',
            'Gerente Regional', 'Gerente', 'Supervisor', 'Dia Todo'
        ]
        ws.append(headers)

        # Dados
        for grupo in grupos:
            ws.append(grupo)

        # Ajusta largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Salva em buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        print(f"✅ Exportação concluída: {len(grupos)} grupos")

        return output

    except Exception as e:
        print(f"❌ Erro na exportação: {str(e)}")
        raise


def validar_planilha(filepath):
    """
    Valida se planilha tem estrutura correta

    Args:
        filepath: Caminho do arquivo Excel

    Returns:
        Tuple (valido: bool, mensagem: str)
    """
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Verifica se tem pelo menos 2 linhas (cabeçalho + 1 linha de dados)
        if ws.max_row < 2:
            return False, "Planilha vazia ou sem dados"

        # Verifica se tem pelo menos 13 colunas
        if ws.max_column < 13:
            return False, "Planilha não tem todas as colunas necessárias"

        return True, "Planilha válida"

    except Exception as e:
        return False, f"Erro ao validar planilha: {str(e)}"
